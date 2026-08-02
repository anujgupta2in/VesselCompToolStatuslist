import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
import re
import os
from io import BytesIO


def extract_date_from_filename(filename):
    base_name = os.path.splitext(os.path.basename(filename))[0]
    date_part = base_name.split()[-1]
    return f"{date_part[0:2]}-{date_part[2:4]}-{date_part[4:]}"


def get_vessel_name(df):
    if "Vessel" in df.columns:
        vessel = df["Vessel"].dropna().astype(str).iloc[0].strip()
        return vessel
    return "Unknown Vessel"


def rename_machinery(value):
    original_value = str(value).strip()
    original_value = re.sub(r"\s+", " ", original_value)
    original_value = re.sub(r"[–—]", "-", original_value)
    compact_value = re.sub(r"\s+", "", original_value)

    specific_mapping = {
        r"^Provision CraneA-?P$": "Provision Crane A-P",
        r"^Provision CraneAft-?Port$": "Provision Crane A-P",
        r"^Provision CraneF-?P$": "Provision Crane F-P",
        r"^Provision CraneF-?S$": "Provision Crane F-S",
        r"^Provision CraneFwd-?P$": "Provision Crane F-P",
        r"^Provision CraneFwd-?Port$": "Provision Crane F-P",
        r"^Provision CraneFwd-?Stbd$": "Provision Crane F-S",
        r"^Provision Crane F-S$": "Provision Crane F-S",
        r"^Provision CraneP1$": "Provision Crane P1",
        r"^Provision CranePort1$": "Provision Crane P1",
        r"^Provision CraneS1$": "Provision Crane S1",
        r"^Provision CraneStarboard1$": "Provision Crane S1",
        r"^Liferaft/Rescue Boat DavitS$": "Liferaft/Rescue Boat Davit S",
        r"^Liferaft/Rescue Boat DavitStarboard$": "Liferaft/Rescue Boat Davit S",
        r"^Rescue BoatS$": "Rescue Boat S",
        r"^Rescue BoatStarboard$": "Rescue Boat S",
        r"^Chain LockerP1$": "Chain Locker P1",
        r"^Chain LockerPort1$": "Chain Locker P1",
        r"^Chain LockerS1$": "Chain Locker S1",
        r"^Chain LockerStarboard1$": "Chain Locker S1",
        r"^Combined Windlass Mooring WinchF1$": "Combined Windlass Mooring Winch F1",
        r"^Combined Windlass Mooring WinchF2$": "Combined Windlass Mooring Winch F2",
        r"^Combined Windlass Mooring WinchForward1$": "Combined Windlass Mooring Winch F1",
        r"^Combined Windlass Mooring WinchForward2$": "Combined Windlass Mooring Winch F2",
        r"^Mooring WinchA1$": "Mooring Winch A1",
        r"^Mooring WinchA2$": "Mooring Winch A2",
        r"^Mooring WinchAft1$": "Mooring Winch A1",
        r"^Mooring WinchAft2$": "Mooring Winch A2",
        r"^Muster StationA1$": "Muster Station A1",
        r"^Muster StationAft1$": "Muster Station A1",
        r"^Accommodation LadderP1$": "Accommodation Ladder P1",
        r"^Accommodation LadderPort1$": "Accommodation Ladder P1",
        r"^Accommodation LadderS1$": "Accommodation Ladder S1",
        r"^Accommodation LadderStarboard1$": "Accommodation Ladder S1",
        r"^Anchor Chain CableP1$": "Anchor Chain Cable P1",
        r"^Anchor Chain CablePort1$": "Anchor Chain Cable P1",
        r"^Anchor Chain CableS1$": "Anchor Chain Cable S1",
        r"^Anchor Chain CableStarboard1$": "Anchor Chain Cable S1",
        r"^AnchorP1$": "Anchor P1",
        r"^AnchorPort1$": "Anchor P1",
        r"^AnchorS1$": "Anchor S1",
        r"^AnchorStarboard1$": "Anchor S1",
        r"^Pilot Combination LadderP1$": "Pilot Combination Ladder P1",
        r"^Pilot Combination LadderPort1$": "Pilot Combination Ladder P1",
        r"^Pilot Combination LadderS1$": "Pilot Combination Ladder S1",
        r"^Pilot Combination LadderStarboard1$": "Pilot Combination Ladder S1",
        r"^Bunker DavitP1$": "Bunker Davit P1",
        r"^Bunker DavitPort1$": "Bunker Davit P1",
        r"^Bunker DavitS1$": "Bunker Davit S1",
        r"^Bunker DavitStarboard1$": "Bunker Davit S1",
        r"^Combined Windlass Mooring WinchP1$": "Combined Windlass Mooring Winch P1",
        r"^Combined Windlass Mooring WinchPort1$": "Combined Windlass Mooring Winch P1",
        r"^Combined Windlass Mooring WinchS1$": "Combined Windlass Mooring Winch S1",
        r"^Combined Windlass Mooring WinchStarboard1$": "Combined Windlass Mooring Winch S1",
        r"^Pilot Ladder DavitP1$": "Pilot Ladder Davit P1",
        r"^Pilot Ladder DavitPort1$": "Pilot Ladder Davit P1",
        r"^Pilot Ladder DavitS2$": "Pilot Ladder Davit S1",
        r"^Pilot Ladder DavitStarboard2$": "Pilot Ladder Davit S1",
        r"^Seaway EquipmentP1$": "Seaway Equipment P1",
        r"^Seaway EquipmentPort1$": "Seaway Equipment P1",
        r"^Seaway EquipmentS1$": "Seaway Equipment S1",
        r"^Seaway EquipmentStarboard1$": "Seaway Equipment S1",
        r"^LifeboatA1$": "Lifeboat A1",
        r"^LifeboatAft1$": "Lifeboat A1",
        r"^Liferaft Embarkation LadderF1$": "Liferaft Embarkation Ladder F1",
        r"^Liferaft Embarkation LadderForward1$": "Liferaft Embarkation Ladder F1",
        r"^Liferaft Embarkation LadderP1$": "Liferaft Embarkation Ladder P1",
        r"^Liferaft Embarkation LadderPort1$": "Liferaft Embarkation Ladder P1",
        r"^Liferaft Embarkation LadderS1$": "Liferaft Embarkation Ladder S1",
        r"^Liferaft Embarkation LadderStarboard1$": "Liferaft Embarkation Ladder S1",
        r"^LiferaftP1$": "Liferaft P1",
        r"^LiferaftPort1$": "Liferaft P1",
        r"^LiferaftP2$": "Liferaft P2",
        r"^LiferaftPort2$": "Liferaft P2",
        r"^LiferaftS1$": "Liferaft S1",
        r"^LiferaftStarboard1$": "Liferaft S1",
        r"^LiferaftS2$": "Liferaft S2",
        r"^LiferaftStarboard2$": "Liferaft S2",
        r"^Mooring WinchA3$": "Mooring Winch A3",
        r"^Mooring WinchAft3$": "Mooring Winch A3",
        r"^Mooring WinchA4$": "Mooring Winch A4",
        r"^Mooring WinchAft4$": "Mooring Winch A4",
        r"^Mooring WinchF1$": "Mooring Winch F1",
        r"^Mooring WinchForward1$": "Mooring Winch F1",
        r"^Mooring WinchF2$": "Mooring Winch F2",
        r"^Mooring WinchForward2$": "Mooring Winch F2",
        r"^Pilot LadderP1$": "Pilot Ladder P1",
        r"^Pilot LadderPort1$": "Pilot Ladder P1",
        r"^Pilot LadderS1$": "Pilot Ladder S1",
        r"^Pilot LadderStarboard1$": "Pilot Ladder S1",
        r"^Rescue BoatP1$": "Rescue Boat P1",
        r"^Rescue BoatPort1$": "Rescue Boat P1",
        r"^Combined Mooring Winch Hydraulic UnitF1$": "Combined Mooring Winch Hydraulic Unit F1",
        r"^Combined Mooring Winch Hydraulic UnitForward1$": "Combined Mooring Winch Hydraulic Unit F1",
        r"^Emergency Towing SystemA1$": "Emergency Towing System A1",
        r"^Emergency Towing SystemAft1$": "Emergency Towing System A1",
        r"^Emergency Towing SystemF1$": "Emergency Towing System F1",
        r"^Emergency Towing SystemForward1$": "Emergency Towing System F1",
        r"^Liferaft 15P1$": "Liferaft 15P1",
        r"^Liferaft 15P2$": "Liferaft 15P2",
        r"^Liferaft 15Port1$": "Liferaft 15P1",
        r"^Liferaft 15Port2$": "Liferaft 15P2",
        r"^Liferaft 6PF-P1$": "Liferaft 6PF-P1",
        r"^Liferaft 6PFwd-Port1$": "Liferaft 6PF-P1",
        r"^Liferaft Embarkation LadderF-P1$": "Liferaft Embarkation Ladder F-P1",
        r"^Liferaft Embarkation LadderF-S1$": "Liferaft Embarkation Ladder F-S1",
        r"^Liferaft Embarkation LadderFwd-Port1$": "Liferaft Embarkation Ladder F-P1",
        r"^Liferaft Embarkation LadderFwd-Stbd1$": "Liferaft Embarkation Ladder F-S1",
        r"^Mooring Winch Hydraulic UnitA1$": "Mooring Winch Hydraulic Unit A1",
        r"^Mooring Winch Hydraulic UnitAft1$": "Mooring Winch Hydraulic Unit A1",
        r"^Rescue BoatS1$": "Rescue Boat S1",
        r"^Rescue BoatStarboard1$": "Rescue Boat S1",
        r"^SARTP1$": "SART P1",
        r"^SARTPort1$": "SART P1",
        r"^SARTS1$": "SART S1",
        r"^SARTStarboard1$": "SART S1",
        r"^Liferaft 15PPort1$": "Liferaft 15PP1",
        r"^Liferaft 15PPort2$": "Liferaft 15PP2",
        r"^ICCPA1$": "ICCP A1",
        r"^ICCPAft1$": "ICCP A1",
        r"^ICCPF1$": "ICCP F1",
        r"^ICCPForward1$": "ICCP F1",
        r"^Slewing Fuel Hose CraneP1$": "Slewing Fuel Hose Crane P1",
        r"^Slewing Fuel Hose CranePort1$": "Slewing Fuel Hose Crane P1",
        r"^Slewing Fuel Hose CraneS1$": "Slewing Fuel Hose Crane S1",
        r"^Slewing Fuel Hose CraneStarboard1$": "Slewing Fuel Hose Crane S1",
        r"^Combined Windlass Mooring WinchF-P1$": "Combined Windlass Mooring Winch F-P1",
        r"^Combined Windlass Mooring WinchF-S1$": "Combined Windlass Mooring Winch F-S1",
        r"^Combined Windlass Mooring WinchFwd-Port1$": "Combined Windlass Mooring Winch F-P1",
        r"^Combined Windlass Mooring WinchFwd-Stbd1$": "Combined Windlass Mooring Winch F-S1",
        r"^Lifeboat DavitP1$": "Lifeboat Davit P1",
        r"^Lifeboat DavitPort1$": "Lifeboat Davit P1",
        r"^LifeboatP1$": "Lifeboat P1",
        r"^LifeboatPort1$": "Lifeboat P1",
        r"^Liferaft Embarkation LadderP2$": "Liferaft Embarkation Ladder P2",
        r"^Liferaft Embarkation LadderPort2$": "Liferaft Embarkation Ladder P2",
        r"^Liferaft Embarkation LadderS2$": "Liferaft Embarkation Ladder S2",
        r"^Liferaft Embarkation LadderStarboard2$": "Liferaft Embarkation Ladder S2",
        r"^Liferaft/Rescue Boat DavitS1$": "Liferaft/Rescue Boat Davit S1",
        r"^Liferaft/Rescue Boat DavitStarboard1$": "Liferaft/Rescue Boat Davit S1",
        r"^Mooring WinchC1$": "Mooring Winch C1",
        r"^Mooring WinchCentre1$": "Mooring Winch C1",
        r"^Hatch CoverA1$": "Hatch Cover A1",
        r"^Hatch CoverA2$": "Hatch Cover A2",
        r"^Hatch CoverA3$": "Hatch Cover A3",
        r"^Hatch CoverA4$": "Hatch Cover A4",
        r"^Hatch CoverA5$": "Hatch Cover A5",
        r"^Hatch CoverA6$": "Hatch Cover A6",
        r"^Hatch CoverA7$": "Hatch Cover A7",
        r"^Hatch CoverAft1$": "Hatch Cover A1",
        r"^Hatch CoverAft2$": "Hatch Cover A2",
        r"^Hatch CoverAft3$": "Hatch Cover A3",
        r"^Hatch CoverAft4$": "Hatch Cover A4",
        r"^Hatch CoverAft5$": "Hatch Cover A5",
        r"^Hatch CoverAft6$": "Hatch Cover A6",
        r"^Hatch CoverAft7$": "Hatch Cover A7",
        r"^Hatch CoverC1$": "Hatch Cover C1",
        r"^Hatch CoverC2$": "Hatch Cover C2",
        r"^Hatch CoverCentre1$": "Hatch Cover C1",
        r"^Hatch CoverCentre2$": "Hatch Cover C2",
        r"^Hatch CoverF1$": "Hatch Cover F1",
        r"^Hatch CoverF2$": "Hatch Cover F2",
        r"^Hatch CoverF3$": "Hatch Cover F3",
        r"^Hatch CoverF4$": "Hatch Cover F4",
        r"^Hatch CoverF5$": "Hatch Cover F5",
        r"^Hatch CoverF6$": "Hatch Cover F6",
        r"^Hatch CoverF7$": "Hatch Cover F7",
        r"^Hatch CoverForward1$": "Hatch Cover F1",
        r"^Hatch CoverForward2$": "Hatch Cover F2",
        r"^Hatch CoverForward3$": "Hatch Cover F3",
        r"^Hatch CoverForward4$": "Hatch Cover F4",
        r"^Hatch CoverForward5$": "Hatch Cover F5",
        r"^Hatch CoverForward6$": "Hatch Cover F6",
        r"^Hatch CoverForward7$": "Hatch Cover F7",
        r"^Mooring WinchC2$": "Mooring Winch C2",
        r"^Mooring WinchCentre2$": "Mooring Winch C2",
        r"^Mooring WinchP1$": "Mooring Winch P1",
        r"^Mooring WinchP2$": "Mooring Winch P2",
        r"^Mooring WinchP3$": "Mooring Winch P3",
        r"^Mooring WinchPort1$": "Mooring Winch P1",
        r"^Mooring WinchPort2$": "Mooring Winch P2",
        r"^Mooring WinchPort3$": "Mooring Winch P3",
        r"^Mooring WinchS1$": "Mooring Winch S1",
        r"^Mooring WinchS2$": "Mooring Winch S2",
        r"^Mooring WinchStarboard1$": "Mooring Winch S1",
        r"^Mooring WinchStarboard2$": "Mooring Winch S2",
        r"^Lifeboat/Rescue BoatS1$": "Lifeboat/Rescue Boat S1",
        r"^Lifeboat/Rescue BoatStarboard1$": "Lifeboat/Rescue Boat S1",
        r"^LiferaftF1$": "Liferaft F1",
        r"^LiferaftForward1$": "Liferaft F1",
        r"^Muster StationP1$": "Muster Station P1",
        r"^Muster StationPort1$": "Muster Station P1",
        r"^Muster StationS1$": "Muster Station S1",
        r"^Muster StationStarboard1$": "Muster Station S1",
        r"^Pilot Combination LadderP2$": "Pilot Combination Ladder P2",
        r"^Pilot Combination LadderPort2$": "Pilot Combination Ladder P2",
        r"^LiferaftFP$": "Liferaft FP",
        r"^LiferaftFS$": "Liferaft FS",
        r"^LiferaftFwd-P$": "Liferaft FP",
        r"^LiferaftFwdS$": "Liferaft FS",
        r"^Lifeboat DavitS1$": "Lifeboat Davit S1",
        r"^Lifeboat DavitStarboard1$": "Lifeboat Davit S1",
        r"^Lifeboat/Rescue BoatP1$": "Lifeboat/Rescue Boat P1",
        r"^Lifeboat/Rescue BoatPort1$": "Lifeboat/Rescue Boat P1",
        r"^LifeboatS1$": "Lifeboat S1",
        r"^LifeboatStarboard1$": "Lifeboat S1",
        r"^Liferaft 16 PersonP1$": "Liferaft 16 Person P1",
        r"^Liferaft 16 PersonP2$": "Liferaft 16 Person P2",
        r"^Liferaft 16 PersonPort1$": "Liferaft 16 Person P1",
        r"^Liferaft 16 PersonPort2$": "Liferaft 16 Person P2",
        r"^Liferaft 16 PersonS1$": "Liferaft 16 Person S1",
        r"^Liferaft 16 PersonS2$": "Liferaft 16 Person S2",
        r"^Liferaft 16 PersonStarboard1$": "Liferaft 16 Person S1",
        r"^Liferaft 16 PersonStarboard2$": "Liferaft 16 Person S2",
        r"^Liferaft 6 PersonF-P1$": "Liferaft 6 Person F-P1",
        r"^Liferaft 6 PersonFwd-Port1$": "Liferaft 6 Person F-P1",
        r"^Liferaft/Rescue Boat DavitP1$": "Liferaft/Rescue Boat Davit P1",
        r"^Liferaft/Rescue Boat DavitPort1$": "Liferaft/Rescue Boat Davit P1",
        r"^Mooring WinchM1$": "Mooring Winch M1",
        r"^Mooring WinchM2$": "Mooring Winch M2",
        r"^Mooring WinchM3$": "Mooring Winch M3",
        r"^Mooring WinchM4$": "Mooring Winch M4",
        r"^Mooring WinchM5$": "Mooring Winch M5",
        r"^Mooring WinchM6$": "Mooring Winch M6",
        r"^Mooring WinchMiddle1$": "Mooring Winch M1",
        r"^Mooring WinchMiddle2$": "Mooring Winch M2",
        r"^Mooring WinchMiddle3$": "Mooring Winch M3",
        r"^Mooring WinchMiddle4$": "Mooring Winch M4",
        r"^Mooring WinchMiddle5$": "Mooring Winch M5",
        r"^Mooring WinchMiddle6$": "Mooring Winch M6",
        r"^Liferaft/Rescue Boat DavitS2$": "Liferaft/Rescue Boat Davit S2",
        r"^Liferaft/Rescue Boat DavitStarboard2$": "Liferaft/Rescue Boat Davit S2",
        r"^Lifeboat/Rescue Boat DavitS1$": "Lifeboat/Rescue Boat Davit S1",
        r"^Lifeboat/Rescue Boat DavitStarboard1$": "Lifeboat/Rescue Boat Davit S1",
        r"^Liferaft Embarkation LadderP3$": "Liferaft Embarkation Ladder P3",
        r"^Liferaft Embarkation LadderPort3$": "Liferaft Embarkation Ladder P3",
        r"^Liferaft Embarkation LadderS3$": "Liferaft Embarkation Ladder S3",
        r"^Liferaft Embarkation LadderStarboard3$": "Liferaft Embarkation Ladder S3",
        r"^Liferaft 6 PersonF1$": "Liferaft 6 Person F1",
        r"^Liferaft 6 PersonForward1$": "Liferaft 6 Person F1",
        r"^Mooring WinchA-P1$": "Mooring Winch A-P1",
        r"^Mooring WinchA-P2$": "Mooring Winch A-P2",
        r"^Mooring WinchA-S1$": "Mooring Winch A-S1",
        r"^Mooring WinchA-S2$": "Mooring Winch A-S2",
        r"^Mooring WinchAft-Port1$": "Mooring Winch A-P1",
        r"^Mooring WinchAft-Port2$": "Mooring Winch A-P2",
        r"^Mooring WinchAft-Stbd1$": "Mooring Winch A-S1",
        r"^Mooring WinchAft-Stbd2$": "Mooring Winch A-S2",
        r"^Mooring WinchF-P1$": "Mooring Winch F-P1",
        r"^Mooring WinchF-S1$": "Mooring Winch F-S1",
        r"^Mooring WinchFwd-Port1$": "Mooring Winch F-P1",
        r"^Mooring WinchFwd-Stbd1$": "Mooring Winch F-S1",
        r"^Combined Mooring Winch Hydraulic UnitA1$": "Combined Mooring Winch Hydraulic Unit A1",
        r"^Combined Mooring Winch Hydraulic UnitAft1$": "Combined Mooring Winch Hydraulic Unit A1",
        r"^Emergency Towing SystemF2$": "Emergency Towing System F2",
        r"^Emergency Towing SystemForward2$": "Emergency Towing System F2",
        r"^Liferaft 20 PersonP1$": "Liferaft 20 Person P1",
        r"^Liferaft 20 PersonP2$": "Liferaft 20 Person P2",
        r"^Liferaft 20 PersonPort1$": "Liferaft 20 Person P1",
        r"^Liferaft 20 PersonPort2$": "Liferaft 20 Person P2",
        r"^Liferaft 20 PersonS1$": "Liferaft 20 Person S1",
        r"^Liferaft 20 PersonS2$": "Liferaft 20 Person S2",
        r"^Liferaft 20 PersonStarboard1$": "Liferaft 20 Person S1",
        r"^Liferaft 20 PersonStarboard2$": "Liferaft 20 Person S2",
        r"^Mooring Winch Hydraulic UnitF1$": "Mooring Winch Hydraulic Unit F1",
        r"^Mooring Winch Hydraulic UnitForward1$": "Mooring Winch Hydraulic Unit F1",
        r"^Provision Crane StbdS1$": "Provision Crane S1",
        r"^Provision Crane StbdStarboard1$": "Provision Crane S1",
        r"^Liferaft Embarkation LadderFS$": "Liferaft Embarkation Ladder FS",
        r"^Liferaft Embarkation LadderFwdS$": "Liferaft Embarkation Ladder FS",
        r"^Combined Mooring Winch Hydraulic UnitF2$": "Combined Mooring Winch Hydraulic Unit F2",
        r"^Combined Mooring Winch Hydraulic UnitForward2$": "Combined Mooring Winch Hydraulic Unit F2",
        r"^Mooring Winch Hydraulic UnitA2$": "Mooring Winch Hydraulic Unit A2",
        r"^Mooring Winch Hydraulic UnitAft2$": "Mooring Winch Hydraulic Unit A2",
        r"^Combined Windlass Mooring WinchFP$": "Combined Windlass Mooring Winch FP",
        r"^Combined Windlass Mooring WinchFS$": "Combined Windlass Mooring Winch FS",
        r"^Combined Windlass Mooring WinchFwd-P$": "Combined Windlass Mooring Winch FP",
        r"^Combined Windlass Mooring WinchFwdS$": "Combined Windlass Mooring Winch FS",
        r"^Mooring WinchA-P3$": "Mooring Winch A-P3",
        r"^Mooring WinchA-S4$": "Mooring Winch A-S3",
        r"^Mooring WinchAft-Port3$": "Mooring Winch A-P3",
        r"^Mooring WinchAft-Stbd4$": "Mooring Winch A-S3",
        r"^Liferaft 15 PersonS1$": "Liferaft 15 Person S1",
        r"^Liferaft 15 PersonS2$": "Liferaft 15 Person S2",
        r"^Liferaft 15 PersonStarboard1$": "Liferaft 15 Person S1",
        r"^Liferaft 15 PersonStarboard2$": "Liferaft 15 Person S2",
        r"^Bilge WellC1$": "Bilge Well C1",
        r"^Bilge WellCentre1$": "Bilge Well C1",
        r"^Bilge WellP1$": "Bilge Well P1",
        r"^Bilge WellPort1$": "Bilge Well P1",
        r"^Bilge WellS1$": "Bilge Well S1",
        r"^Bilge WellStarboard1$": "Bilge Well S1",
        r"^Chain LockerC1$": "Chain Locker C1",
        r"^Chain LockerC2$": "Chain Locker C2",
        r"^Chain LockerCentre1$": "Chain Locker C1",
        r"^Chain LockerCentre2$": "Chain Locker C2",
        r"^Suez Search Light DavitF1$": "Suez Search Light Davit F1",
        r"^Suez Search Light DavitForward1$": "Suez Search Light Davit F1",
        r"^Liferaft 15 PersonP1$": "Liferaft 15 Person P1",
        r"^Liferaft 15 PersonP2$": "Liferaft 15 Person P2",
        r"^Liferaft 15 PersonPort1$": "Liferaft 15 Person P1",
        r"^Liferaft 15 PersonPort2$": "Liferaft 15 Person P2",
        r"^Liferaft 6 PersonC1$": "Liferaft 6 Person C1",
        r"^Liferaft 6 PersonCentre1$": "Liferaft 6 Person C1",
        r"^Lifeboat Davit\.S1$": "Lifeboat Davit S",
        r"^Lifeboat Davit\.Starboard1$": "Lifeboat Davit S",
    }

    for pattern, replacement in specific_mapping.items():
        if re.match(pattern, original_value, flags=re.IGNORECASE):
            return replacement

    suffix_mapping = {
        r"(.*)(?:Aft)$": r"\1A",
        r"(.*)(?:Forward)$": r"\1F",
        r"(.*)(?:Fwd)$": r"\1F",
        r"(.*)(?:Port)$": r"\1P",
        r"(.*)(?:Starboard)$": r"\1S",
        r"(.*)(?:-P)$": r"\1P",
        r"(.*)(?:-S)$": r"\1S",
        r"(.*)(?:-Port)$": r"\1P",
        r"(.*)(?:-Stbd)$": r"\1S",
    }

    for pattern, replacement in suffix_mapping.items():
        if re.match(pattern, original_value, flags=re.IGNORECASE):
            return re.sub(pattern, replacement, original_value, flags=re.IGNORECASE).strip()

    return original_value


def _detect_col(df, candidates):
    """Return the first column name from candidates that exists in df, or None."""
    for c in candidates:
        if c in df.columns:
            return c
    return None


def _build_job_detail(df, machinery_col, title_col, code_col, freq_col=None):
    """
    For each machinery, return a DataFrame with columns:
      Job Code | Job Title | Frequency | Count
    where Count is how many rows share that (code, title, frequency) combination.
    Returns a dict: {machinery_name -> DataFrame}
    """
    detail = {}
    keep = [machinery_col]
    if code_col:
        keep.append(code_col)
    if title_col:
        keep.append(title_col)
    if freq_col and freq_col in df.columns:
        keep.append(freq_col)

    sub = df[keep].copy()
    sub[machinery_col] = sub[machinery_col].astype(str)

    for machinery, grp in sub.groupby(machinery_col):
        grp = grp.drop(columns=[machinery_col])
        if grp.empty:
            detail[machinery] = pd.DataFrame(columns=['Job Code', 'Job Title', 'Frequency', 'Count'])
            continue

        rename_map = {}
        if code_col and code_col in grp.columns:
            rename_map[code_col] = 'Job Code'
        if title_col and title_col in grp.columns:
            rename_map[title_col] = 'Job Title'
        if freq_col and freq_col in grp.columns:
            rename_map[freq_col] = 'Frequency'
        grp = grp.rename(columns=rename_map)

        if 'Job Code' not in grp.columns:
            grp['Job Code'] = '-'
        if 'Job Title' not in grp.columns:
            grp['Job Title'] = '-'
        if 'Frequency' not in grp.columns:
            grp['Frequency'] = '-'

        grp['Job Code']   = grp['Job Code'].fillna('-').astype(str)
        grp['Job Title']  = grp['Job Title'].fillna('-').astype(str)
        grp['Frequency']  = grp['Frequency'].fillna('-').astype(str)

        counted = (
            grp.groupby(['Job Code', 'Job Title', 'Frequency'], dropna=False)
            .size()
            .reset_index(name='Count')
            .sort_values(['Job Code', 'Job Title', 'Frequency'])
            .reset_index(drop=True)
        )
        detail[machinery] = counted

    return detail


def build_frequency_comparison(detail1, detail2, col1_label, col2_label, common_titles_map=None):
    """
    Compare the Frequency interval (e.g. "3 Months", "12 Months") for each job code
    between the two files, for common job titles only (when common_titles_map is given).

    Returns (freq_df, excel_bytes)
    freq_df columns:
      Machinery | Job Code | Job Title | Frequency (File1) | Frequency (File2) | Match

    Match values: "✓ Match" | "✗ Differ" | "Only in File 1" | "Only in File 2"
    """
    def _agg_freq(df):
        """Collapse to (Job Code, Job Title) → unique Frequency values joined."""
        if df.empty:
            return pd.DataFrame(columns=['Job Code', 'Job Title', 'Frequency'])
        if 'Frequency' not in df.columns:
            df = df.copy()
            df['Frequency'] = '-'
        return (
            df.groupby(['Job Code', 'Job Title'], dropna=False)['Frequency']
            .apply(lambda x: ', '.join(sorted(set(str(v) for v in x if str(v) not in ('', '-', 'nan')))) or '-')
            .reset_index()
        )

    all_machinery = sorted(set(list(detail1.keys()) + list(detail2.keys())))
    rows = []

    for machinery in all_machinery:
        df1 = detail1.get(machinery, pd.DataFrame(columns=['Job Code', 'Job Title', 'Frequency', 'Count']))
        df2 = detail2.get(machinery, pd.DataFrame(columns=['Job Code', 'Job Title', 'Frequency', 'Count']))

        if common_titles_map is not None:
            common = common_titles_map.get(machinery, set())
            if not common:
                continue
            if not df1.empty:
                df1 = df1[df1['Job Title'].isin(common)].copy()
            if not df2.empty:
                df2 = df2[df2['Job Title'].isin(common)].copy()

        agg1 = _agg_freq(df1).rename(columns={'Frequency': col1_label})
        agg2 = _agg_freq(df2).rename(columns={'Frequency': col2_label})

        if agg1.empty and agg2.empty:
            continue

        merged = pd.merge(agg1, agg2, on=['Job Code', 'Job Title'], how='outer').fillna('-')

        def _match_status(row):
            f1 = str(row[col1_label]).strip()
            f2 = str(row[col2_label]).strip()
            if f1 == '-' and f2 != '-':
                return 'Only in File 2'
            if f2 == '-' and f1 != '-':
                return 'Only in File 1'
            if f1 == f2:
                return '✓ Match'
            return '✗ Differ'

        merged['Match'] = merged.apply(_match_status, axis=1)
        merged.insert(0, 'Machinery', machinery)
        rows.append(merged[['Machinery', 'Job Code', 'Job Title', col1_label, col2_label, 'Match']])

    if rows:
        freq_df = pd.concat(rows, ignore_index=True)
    else:
        freq_df = pd.DataFrame(
            columns=['Machinery', 'Job Code', 'Job Title', col1_label, col2_label, 'Match']
        )

    # ---- Generate Excel ----
    output = BytesIO()
    freq_df.to_excel(output, index=False)
    output.seek(0)

    wb = load_workbook(output)
    sheet = wb.active

    fill_red    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    fill_green  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fill_orange = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    fill_blue   = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    hdr_fill    = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    wrap_align  = Alignment(wrap_text=True, vertical='top')

    for c in range(1, 7):
        cell = sheet.cell(row=1, column=c)
        cell.fill = hdr_fill
        cell.font = Font(bold=True, color="FFFFFF")

    sheet.column_dimensions['A'].width = 35
    sheet.column_dimensions['B'].width = 15
    sheet.column_dimensions['C'].width = 50
    sheet.column_dimensions['D'].width = 18
    sheet.column_dimensions['E'].width = 18
    sheet.column_dimensions['F'].width = 16

    for row in range(2, sheet.max_row + 1):
        match_val = str(sheet.cell(row=row, column=6).value or '')
        if match_val == 'Only in File 1':
            fill = fill_orange
        elif match_val == 'Only in File 2':
            fill = fill_blue
        elif match_val == '✗ Differ':
            fill = fill_red
        else:
            fill = fill_green

        for c in range(1, 7):
            cell = sheet.cell(row=row, column=c)
            cell.fill = fill
            cell.alignment = wrap_align

        if match_val in ('✗ Differ', 'Only in File 1', 'Only in File 2'):
            sheet.cell(row=row, column=6).font = Font(bold=True)

    output_final = BytesIO()
    wb.save(output_final)
    output_final.seek(0)

    return freq_df, output_final.getvalue()


def _get_file_label(filename):
    """Return a human-friendly label from a filename, e.g. 'Harzand 15052026'.
    Splits on underscores/spaces and drops long numeric timestamp segments."""
    base = os.path.splitext(os.path.basename(filename))[0]
    parts = re.split(r'[_\s]+', base)
    parts = [p for p in parts if p and not (p.isdigit() and len(p) > 10)]
    return ' '.join(parts)


def process_files(file1_content, file2_content, file1_name, file2_name):
    df_system_mgmt = pd.read_csv(BytesIO(file1_content))
    df_pms_jobs = pd.read_csv(BytesIO(file2_content))

    col1 = _get_file_label(file1_name)
    col2 = _get_file_label(file2_name)

    if col1 == col2:
        col1 += " [File 1]"
        col2 += " [File 2]"

    possible_machinery_columns = ['Machinery', 'Machinery Location', 'Component Name', 'System Name']
    possible_title_columns = ['Job Title', 'Title']
    possible_code_columns = ['Job Code', 'Code', 'Job No', 'Job No.', 'Job Number', 'JobCode']

    for col in possible_machinery_columns:
        if col in df_system_mgmt.columns:
            df_system_mgmt.rename(columns={col: 'Machinery'}, inplace=True)
            break
    else:
        raise ValueError("No recognized Machinery column in first file.")

    for col in possible_machinery_columns:
        if col in df_pms_jobs.columns:
            df_pms_jobs.rename(columns={col: 'Machinery Location'}, inplace=True)
            break
    else:
        raise ValueError("No recognized Machinery column in second file.")

    df_system_mgmt['Machinery'] = df_system_mgmt['Machinery'].apply(rename_machinery)
    df_pms_jobs['Machinery Location'] = df_pms_jobs['Machinery Location'].apply(rename_machinery)

    def _merge_freq_cols(df):
        """If both 'Frequency' and 'Frequency Type' columns exist, combine them
        into a single 'Frequency' column (e.g. '3 Months' + 'Annual' → '3 Months Annual'),
        then drop 'Frequency Type'."""
        if 'Frequency' in df.columns and 'Frequency Type' in df.columns:
            freq = df['Frequency'].fillna('').astype(str).str.strip()
            ftype = df['Frequency Type'].fillna('').astype(str).str.strip()
            combined = (freq + ' ' + ftype).str.strip()
            df['Frequency'] = combined
            df.drop(columns=['Frequency Type'], inplace=True)
        return df

    df_system_mgmt = _merge_freq_cols(df_system_mgmt)
    df_pms_jobs    = _merge_freq_cols(df_pms_jobs)

    title_col1 = _detect_col(df_system_mgmt, possible_title_columns)
    code_col1  = _detect_col(df_system_mgmt, possible_code_columns)
    freq_col1  = _detect_col(df_system_mgmt, ['Frequency'])
    title_col2 = _detect_col(df_pms_jobs, possible_title_columns)
    code_col2  = _detect_col(df_pms_jobs, possible_code_columns)
    freq_col2  = _detect_col(df_pms_jobs, ['Frequency'])

    detail1 = _build_job_detail(df_system_mgmt, 'Machinery', title_col1, code_col1, freq_col1)
    detail2 = _build_job_detail(df_pms_jobs, 'Machinery Location', title_col2, code_col2, freq_col2)

    freq_df, freq_excel = build_frequency_comparison(detail1, detail2, col1, col2)

    system_mgmt_counts = df_system_mgmt['Machinery'].value_counts().reset_index()
    pms_jobs_counts = df_pms_jobs['Machinery Location'].value_counts().reset_index()

    system_mgmt_counts.columns = ['Machinery', col1]
    pms_jobs_counts.columns = ['Machinery', col2]

    comparison_df = pd.merge(system_mgmt_counts, pms_jobs_counts, on='Machinery', how='outer').fillna(0)

    comparison_df[col1] = comparison_df[col1].astype(int)
    comparison_df[col2] = comparison_df[col2].astype(int)
    comparison_df['Difference'] = comparison_df[col1] - comparison_df[col2]

    total_row = {
        'Machinery': 'TOTAL',
        col1: comparison_df[col1].sum(),
        col2: comparison_df[col2].sum(),
        'Difference': comparison_df['Difference'].sum()
    }
    comparison_df = pd.concat([comparison_df, pd.DataFrame([total_row])], ignore_index=True)

    output = BytesIO()
    comparison_df.to_excel(output, index=False)
    output.seek(0)

    wb = load_workbook(output)
    sheet = wb.active

    fill_red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    fill_green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fill_yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    bold_font = Font(bold=True)
    red_font = Font(color="9C0006")
    green_font = Font(color="006100")

    for row in range(2, sheet.max_row + 1):
        machinery = sheet.cell(row=row, column=1)
        count1 = sheet.cell(row=row, column=2).value
        count2 = sheet.cell(row=row, column=3).value
        diff_cell = sheet.cell(row=row, column=4)

        if machinery.value != 'TOTAL':
            if count1 == 0 or count2 == 0:
                machinery.fill = fill_red
                machinery.font = bold_font
                diff_cell.fill = fill_red
                diff_cell.font = red_font
            if count1 != count2:
                sheet.cell(row=row, column=2).fill = fill_yellow
                sheet.cell(row=row, column=3).fill = fill_yellow
                if count1 > count2:
                    diff_cell.fill = fill_green
                    diff_cell.font = green_font
                else:
                    diff_cell.fill = fill_red
                    diff_cell.font = red_font
        else:
            for c in range(1, 5):
                sheet.cell(row=row, column=c).font = bold_font

    # ---- Job Detail Breakdown sheet ----
    detail_sheet = wb.create_sheet(title="Job Detail Breakdown")

    header_fill   = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font   = Font(bold=True, color="FFFFFF")
    subhdr_fill1  = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    subhdr_fill2  = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
    subhdr_font   = Font(bold=True, color="FFFFFF")
    dup_fill      = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    only1_fill    = PatternFill(start_color="FFD180", end_color="FFD180", fill_type="solid")
    only2_fill    = PatternFill(start_color="BBDEFB", end_color="BBDEFB", fill_type="solid")
    alt_fill      = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    machinery_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    machinery_font = Font(bold=True)
    wrap_align    = Alignment(wrap_text=True, vertical='top')

    diff_machineries = comparison_df[
        (comparison_df['Machinery'] != 'TOTAL') &
        (comparison_df[col1] != comparison_df[col2])
    ]['Machinery'].tolist()

    detail_sheet.column_dimensions['A'].width = 8   # No.
    detail_sheet.column_dimensions['B'].width = 22  # Job Code (file 1)
    detail_sheet.column_dimensions['C'].width = 55  # Job Title (file 1)
    detail_sheet.column_dimensions['D'].width = 8   # Count (file 1)
    detail_sheet.column_dimensions['E'].width = 8   # No. (file 2)
    detail_sheet.column_dimensions['F'].width = 22  # Job Code (file 2)
    detail_sheet.column_dimensions['G'].width = 55  # Job Title (file 2)
    detail_sheet.column_dimensions['H'].width = 8   # Count (file 2)

    # Legend row at the top of the detail sheet
    legend_items = [
        ("Only in left file", "FFD180"),
        ("Only in right file", "BBDEFB"),
        ("Duplicate Job Code", "FFF3CD"),
    ]
    detail_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    legend_hdr = detail_sheet.cell(row=1, column=1, value="Legend:")
    legend_hdr.font = Font(bold=True)
    col_offset = 3
    for label, color in legend_items:
        swatch_cell = detail_sheet.cell(row=1, column=col_offset, value="  " + label)
        swatch_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        swatch_cell.font = Font(bold=False)
        detail_sheet.column_dimensions[
            detail_sheet.cell(row=1, column=col_offset).column_letter
        ].width = max(18, len(label) + 4)
        col_offset += 2

    cur_row = 3  # leave a blank row after the legend

    for machinery in diff_machineries:
        # Machinery name header spanning all columns
        detail_sheet.merge_cells(
            start_row=cur_row, start_column=1,
            end_row=cur_row, end_column=8
        )
        mach_cell = detail_sheet.cell(row=cur_row, column=1, value=machinery)
        mach_cell.fill = machinery_fill
        mach_cell.font = machinery_font
        mach_cell.alignment = Alignment(vertical='center')
        cur_row += 1

        # Sub-header row 1: file labels (merged spans)
        detail_sheet.merge_cells(
            start_row=cur_row, start_column=1, end_row=cur_row, end_column=4
        )
        file1_cell = detail_sheet.cell(row=cur_row, column=1, value=col1)
        file1_cell.fill = subhdr_fill1
        file1_cell.font = subhdr_font
        file1_cell.alignment = Alignment(vertical='center')

        detail_sheet.merge_cells(
            start_row=cur_row, start_column=5, end_row=cur_row, end_column=8
        )
        file2_cell = detail_sheet.cell(row=cur_row, column=5, value=col2)
        file2_cell.fill = subhdr_fill2
        file2_cell.font = subhdr_font
        file2_cell.alignment = Alignment(vertical='center')
        cur_row += 1

        # Sub-header row 2: column labels
        col_headers = [
            (1, 'No.',       subhdr_fill1),
            (2, 'Job Code',  subhdr_fill1),
            (3, 'Job Title', subhdr_fill1),
            (4, 'Count',     subhdr_fill1),
            (5, 'No.',       subhdr_fill2),
            (6, 'Job Code',  subhdr_fill2),
            (7, 'Job Title', subhdr_fill2),
            (8, 'Count',     subhdr_fill2),
        ]
        for c, val, fill in col_headers:
            cell = detail_sheet.cell(row=cur_row, column=c, value=val)
            cell.fill = fill
            cell.font = subhdr_font
            cell.alignment = wrap_align
        cur_row += 1

        df1 = detail1.get(machinery, pd.DataFrame(columns=['Job Code', 'Job Title', 'Count']))
        df2 = detail2.get(machinery, pd.DataFrame(columns=['Job Code', 'Job Title', 'Count']))

        max_rows = max(len(df1), len(df2), 1)
        df1 = df1.reset_index(drop=True)
        df2 = df2.reset_index(drop=True)

        dup_codes1  = set(df1.loc[df1['Count'] > 1, 'Job Code'].astype(str).str.strip()) if not df1.empty else set()
        dup_codes2  = set(df2.loc[df2['Count'] > 1, 'Job Code'].astype(str).str.strip()) if not df2.empty else set()
        all_codes1  = set(df1['Job Code'].astype(str).str.strip()) if not df1.empty else set()
        all_codes2  = set(df2['Job Code'].astype(str).str.strip()) if not df2.empty else set()
        excl_codes1 = all_codes1 - all_codes2   # only in file 1 → orange
        excl_codes2 = all_codes2 - all_codes1   # only in file 2 → blue

        def _row_fill1(code, count, alt):
            c = str(code).strip()
            if c in excl_codes1:
                return only1_fill
            if count > 1:
                return dup_fill
            return alt or PatternFill()

        def _row_fill2(code, count, alt):
            c = str(code).strip()
            if c in excl_codes2:
                return only2_fill
            if count > 1:
                return dup_fill
            return alt or PatternFill()

        for i in range(max_rows):
            row_alt = alt_fill if i % 2 == 1 else None

            # File 1 columns: A=No., B=Job Code, C=Job Title, D=Count (cols 1-4)
            if i < len(df1):
                r1 = df1.iloc[i]
                fill1 = _row_fill1(r1['Job Code'], int(r1['Count']), row_alt)
                detail_sheet.cell(row=cur_row, column=1, value=i + 1).fill = fill1
                for c, val in enumerate([r1['Job Code'], r1['Job Title'], int(r1['Count'])], 2):
                    cell = detail_sheet.cell(row=cur_row, column=c, value=val)
                    cell.fill = fill1
                    cell.alignment = wrap_align
                if int(r1['Count']) > 1:
                    detail_sheet.cell(row=cur_row, column=4).font = Font(bold=True, color="9C6500")
            else:
                for c in range(1, 5):
                    detail_sheet.cell(row=cur_row, column=c, value='').fill = row_alt or PatternFill()

            # File 2 columns: E=No., F=Job Code, G=Job Title, H=Count (cols 5-8)
            if i < len(df2):
                r2 = df2.iloc[i]
                fill2 = _row_fill2(r2['Job Code'], int(r2['Count']), row_alt)
                detail_sheet.cell(row=cur_row, column=5, value=i + 1).fill = fill2
                for c, val in enumerate([r2['Job Code'], r2['Job Title'], int(r2['Count'])], 6):
                    cell = detail_sheet.cell(row=cur_row, column=c, value=val)
                    cell.fill = fill2
                    cell.alignment = wrap_align
                if int(r2['Count']) > 1:
                    detail_sheet.cell(row=cur_row, column=8).font = Font(bold=True, color="9C6500")
            else:
                for c in range(5, 9):
                    detail_sheet.cell(row=cur_row, column=c, value='').fill = row_alt or PatternFill()

            cur_row += 1

        cur_row += 1  # blank separator row

    if not diff_machineries:
        detail_sheet.cell(row=1, column=1, value="No machinery with count differences found.")

    output_final = BytesIO()
    wb.save(output_final)
    output_final.seek(0)

    job_detail = {
        'col1': col1,
        'col2': col2,
        'detail1': detail1,
        'detail2': detail2,
        'freq_df': freq_df,
        'freq_excel': freq_excel,
    }

    return comparison_df, output_final.getvalue(), job_detail


# ---------------------------------------------------------------------------
# Tab 4 – New Jobs Analysis
# ---------------------------------------------------------------------------

def build_new_jobs_analysis(file1_content, file2_content, file1_name, file2_name):
    """Filter BOTH files to rows where Status == 'New', then compare per machinery.

    Summary columns:
        Machinery Location | New in {label1} | New in {label2} |
        Common in Both | Not in {label1} | Not in {label2}

    "New in X"   = total rows (incl. duplicates) with Status=New in that file
    "Common"     = distinct job codes present as New in both files
    "Not in X"   = distinct codes New in the other file but absent from X

    detail_dict: machinery → DataFrame
        Job Code | Job Title | Machinery | File Name | Job Status | Match
    """
    from openpyxl import Workbook

    df1 = pd.read_csv(BytesIO(file1_content))
    df2 = pd.read_csv(BytesIO(file2_content))

    label1 = _get_file_label(file1_name)
    label2 = _get_file_label(file2_name)

    possible_machinery = ['Machinery', 'Machinery Location', 'Component Name', 'System Name']
    possible_title     = ['Job Title', 'Title']
    possible_code      = ['Job Code', 'Code', 'Job No', 'Job No.', 'Job Number', 'JobCode']
    possible_status    = ['Status', 'Job Status']

    for col in possible_machinery:
        if col in df1.columns:
            df1.rename(columns={col: 'Machinery'}, inplace=True)
            break
    else:
        raise ValueError("No recognised Machinery column in first file.")

    for col in possible_machinery:
        if col in df2.columns:
            df2.rename(columns={col: 'Machinery'}, inplace=True)
            break
    else:
        raise ValueError("No recognised Machinery column in second file.")

    df1['Machinery'] = df1['Machinery'].apply(rename_machinery)
    df2['Machinery'] = df2['Machinery'].apply(rename_machinery)

    title_col1 = _detect_col(df1, possible_title)
    code_col1  = _detect_col(df1, possible_code)
    title_col2 = _detect_col(df2, possible_title)
    code_col2  = _detect_col(df2, possible_code)
    status_col1 = _detect_col(df1, possible_status)
    status_col2 = _detect_col(df2, possible_status)

    missing = []
    if not status_col1:
        missing.append(f"first file (expected: {', '.join(possible_status)})")
    if not status_col2:
        missing.append(f"second file (expected: {', '.join(possible_status)})")
    if missing:
        raise ValueError("No Status column found in: " + "; ".join(missing))

    df1_new = df1[df1[status_col1].astype(str).str.strip().str.lower() == 'new'].copy()
    df2_new = df2[df2[status_col2].astype(str).str.strip().str.lower() == 'new'].copy()

    all_machinery = sorted(set(
        df1_new['Machinery'].dropna().astype(str).tolist() +
        df2_new['Machinery'].dropna().astype(str).tolist()
    ))

    col_new1      = f'New in {label1}'
    col_new2      = f'New in {label2}'
    col_common    = 'Common in Both'
    col_not_in1   = f'Not in {label1}'
    col_not_in2   = f'Not in {label2}'

    summary_rows = []
    detail_dict  = {}

    def _pick_title(df, code_col, title_col, code):
        if not code_col:
            return ''
        sub = df[df[code_col].astype(str).str.strip() == code]
        if title_col and not sub.empty:
            return str(sub.iloc[0][title_col]).strip()
        return ''

    for machinery in all_machinery:
        m1 = df1_new[df1_new['Machinery'].astype(str) == machinery]
        m2 = df2_new[df2_new['Machinery'].astype(str) == machinery]

        count1 = len(m1)
        count2 = len(m2)

        codes1 = set(m1[code_col1].astype(str).str.strip()) if code_col1 and not m1.empty else set()
        codes2 = set(m2[code_col2].astype(str).str.strip()) if code_col2 and not m2.empty else set()

        in_both  = codes1 & codes2
        only1    = codes1 - codes2
        only2    = codes2 - codes1

        summary_rows.append({
            'Machinery Location': machinery,
            col_new1:   count1,
            col_new2:   count2,
            col_common: len(in_both),
            col_not_in1: len(only2),   # in file2 New but not file1
            col_not_in2: len(only1),   # in file1 New but not file2
        })

        rows = []
        for code in sorted(in_both):
            title = _pick_title(m1, code_col1, title_col1, code) or \
                    _pick_title(m2, code_col2, title_col2, code)
            rows.append({'Job Code': code, 'Job Title': title,
                         'Machinery': machinery,
                         'File Name': f'{label1} & {label2}',
                         'Job Status': 'New', 'Match': 'In Both'})
        for code in sorted(only1):
            rows.append({'Job Code': code,
                         'Job Title': _pick_title(m1, code_col1, title_col1, code),
                         'Machinery': machinery,
                         'File Name': label1,
                         'Job Status': 'New', 'Match': f'Only in {label1}'})
        for code in sorted(only2):
            rows.append({'Job Code': code,
                         'Job Title': _pick_title(m2, code_col2, title_col2, code),
                         'Machinery': machinery,
                         'File Name': label2,
                         'Job Status': 'New', 'Match': f'Only in {label2}'})

        if rows:
            detail_dict[machinery] = pd.DataFrame(rows)

    summary_df = pd.DataFrame(summary_rows) if summary_rows else pd.DataFrame(
        columns=['Machinery Location', col_new1, col_new2,
                 col_common, col_not_in1, col_not_in2]
    )

    # ---- Excel export ----
    wb = Workbook()
    ws_sum = wb.active
    ws_sum.title = "New Jobs Summary"

    hdr_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF")
    grn_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    ora_fill = PatternFill(start_color="FFD180", end_color="FFD180", fill_type="solid")
    blu_fill = PatternFill(start_color="BBDEFB", end_color="BBDEFB", fill_type="solid")
    alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    wrap_aln = Alignment(wrap_text=True, vertical='top')

    sum_headers = list(summary_df.columns)
    col_widths  = [40, 18, 18, 16, 20, 20]
    for ci, (hdr, w) in enumerate(zip(sum_headers, col_widths), 1):
        cell = ws_sum.cell(row=1, column=ci, value=hdr)
        cell.fill = hdr_fill
        cell.font = hdr_font
        ws_sum.column_dimensions[cell.column_letter].width = w

    for ri, row_data in enumerate(summary_df.itertuples(index=False), 2):
        base = alt_fill if ri % 2 == 0 else PatternFill()
        for ci, val in enumerate(row_data, 1):
            ws_sum.cell(row=ri, column=ci, value=val).fill = base
        v_common = row_data[3]
        v_not1   = row_data[4]
        v_not2   = row_data[5]
        if v_common > 0:
            ws_sum.cell(row=ri, column=4).fill = grn_fill
        if v_not1 > 0:
            ws_sum.cell(row=ri, column=5).fill = ora_fill
        if v_not2 > 0:
            ws_sum.cell(row=ri, column=6).fill = blu_fill

    # Sheet 2: Detail
    ws_det = wb.create_sheet(title="New Jobs Detail")
    mach_fill    = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    det_hdr_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    det_cols     = ['Job Code', 'Job Title', 'Machinery', 'File Name', 'Job Status', 'Match']
    det_widths   = [22, 55, 35, 28, 12, 28]

    for ci, (hdr, w) in enumerate(zip(det_cols, det_widths), 1):
        cell = ws_det.cell(row=1, column=ci, value=hdr)
        cell.fill = det_hdr_fill
        cell.font = Font(bold=True, color="FFFFFF")
        ws_det.column_dimensions[cell.column_letter].width = w

    det_cur = 2
    for _, df_det in detail_dict.items():
        for row_det in df_det.itertuples(index=False):
            match = str(row_det[5])
            rfill = grn_fill if match == 'In Both' else (
                    ora_fill if match.startswith('Only in ' + label1) else blu_fill)
            for ci, val in enumerate(row_det, 1):
                cell = ws_det.cell(row=det_cur, column=ci, value=val)
                cell.fill = rfill
                cell.alignment = wrap_aln
            det_cur += 1

    if not detail_dict:
        ws_det.cell(row=2, column=1, value="No 'New' status jobs found in either file.")

    out = BytesIO()
    wb.save(out)
    out.seek(0)

    return summary_df, detail_dict, out.getvalue(), label1, label2


# ---------------------------------------------------------------------------
# Tab 5 – Critical Jobs Analysis
# ---------------------------------------------------------------------------

def build_critical_jobs_analysis(file1_content, file2_content, file1_name, file2_name):
    """Filter BOTH files to rows where Column B (2nd column) == 'C' (critical),
    then compare per machinery location.

    Summary columns:
        Machinery Location | Critical in {label1} | Critical in {label2} |
        Common in Both | Not in {label1} | Not in {label2}

    detail_dict: machinery → DataFrame
        Job Code | Job Title | Machinery | File Name | Match
    """
    from openpyxl import Workbook

    df1 = pd.read_csv(BytesIO(file1_content))
    df2 = pd.read_csv(BytesIO(file2_content))

    label1 = _get_file_label(file1_name)
    label2 = _get_file_label(file2_name)

    def _filter_critical(df):
        """Return only rows where the 2nd column value is 'C'."""
        if df.shape[1] < 2:
            return df.iloc[0:0].copy()
        col_b = df.columns[1]
        return df[df[col_b].astype(str).str.strip().str.upper() == 'C'].copy()

    df1_crit = _filter_critical(df1)
    df2_crit = _filter_critical(df2)

    possible_machinery = ['Machinery', 'Machinery Location', 'Component Name', 'System Name']
    possible_title     = ['Job Title', 'Title']
    possible_code      = ['Job Code', 'Code', 'Job No', 'Job No.', 'Job Number', 'JobCode']

    for col in possible_machinery:
        if col in df1_crit.columns:
            df1_crit = df1_crit.rename(columns={col: 'Machinery'})
            break
    else:
        raise ValueError("No recognised Machinery column in first file.")

    for col in possible_machinery:
        if col in df2_crit.columns:
            df2_crit = df2_crit.rename(columns={col: 'Machinery'})
            break
    else:
        raise ValueError("No recognised Machinery column in second file.")

    df1_crit['Machinery'] = df1_crit['Machinery'].apply(rename_machinery)
    df2_crit['Machinery'] = df2_crit['Machinery'].apply(rename_machinery)

    title_col1 = _detect_col(df1_crit, possible_title)
    code_col1  = _detect_col(df1_crit, possible_code)
    title_col2 = _detect_col(df2_crit, possible_title)
    code_col2  = _detect_col(df2_crit, possible_code)

    col_crit1    = f'Critical in {label1}'
    col_crit2    = f'Critical in {label2}'
    col_common   = 'Common in Both'
    col_not_in1  = f'Not in {label1}'
    col_not_in2  = f'Not in {label2}'

    all_machinery = sorted(set(
        df1_crit['Machinery'].dropna().astype(str).tolist() +
        df2_crit['Machinery'].dropna().astype(str).tolist()
    ))

    summary_rows = []
    detail_dict  = {}

    def _pick_title(df, code_col, title_col, code):
        if not code_col:
            return ''
        sub = df[df[code_col].astype(str).str.strip() == code]
        if title_col and not sub.empty:
            return str(sub.iloc[0][title_col]).strip()
        return ''

    for machinery in all_machinery:
        m1 = df1_crit[df1_crit['Machinery'].astype(str) == machinery]
        m2 = df2_crit[df2_crit['Machinery'].astype(str) == machinery]

        count1 = len(m1)
        count2 = len(m2)

        codes1 = set(m1[code_col1].astype(str).str.strip()) if code_col1 and not m1.empty else set()
        codes2 = set(m2[code_col2].astype(str).str.strip()) if code_col2 and not m2.empty else set()

        in_both = codes1 & codes2
        only1   = codes1 - codes2
        only2   = codes2 - codes1

        summary_rows.append({
            'Machinery Location': machinery,
            col_crit1:    count1,
            col_crit2:    count2,
            col_common:   len(in_both),
            col_not_in1:  len(only2),
            col_not_in2:  len(only1),
        })

        detail_rows = []
        for code in sorted(in_both):
            t1 = _pick_title(m1, code_col1, title_col1, code)
            t2 = _pick_title(m2, code_col2, title_col2, code)
            title = t1 or t2
            detail_rows.append({
                'Job Code': code, 'Job Title': title,
                'Machinery': machinery,
                'File Name': f'{label1} & {label2}',
                'Match': 'In Both',
            })
        for code in sorted(only1):
            title = _pick_title(m1, code_col1, title_col1, code)
            detail_rows.append({
                'Job Code': code, 'Job Title': title,
                'Machinery': machinery,
                'File Name': label1,
                'Match': f'Only in {label1}',
            })
        for code in sorted(only2):
            title = _pick_title(m2, code_col2, title_col2, code)
            detail_rows.append({
                'Job Code': code, 'Job Title': title,
                'Machinery': machinery,
                'File Name': label2,
                'Match': f'Only in {label2}',
            })

        if detail_rows:
            detail_dict[machinery] = pd.DataFrame(detail_rows)

    summary_df = pd.DataFrame(summary_rows, columns=[
        'Machinery Location', col_crit1, col_crit2,
        col_common, col_not_in1, col_not_in2,
    ])

    # ── Excel export ─────────────────────────────────────────────────────────
    wb = Workbook()

    grn_fill     = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    ora_fill     = PatternFill(start_color="FFD180", end_color="FFD180", fill_type="solid")
    blu_fill     = PatternFill(start_color="BBDEFB", end_color="BBDEFB", fill_type="solid")
    hdr_fill     = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    det_hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    wrap_aln     = Alignment(wrap_text=True, vertical='top')

    # Summary sheet
    ws_sum = wb.active
    ws_sum.title = "Critical Jobs Summary"
    sum_cols = list(summary_df.columns)
    for ci, hdr in enumerate(sum_cols, 1):
        cell = ws_sum.cell(row=1, column=ci, value=hdr)
        cell.fill = hdr_fill
        cell.font = Font(bold=True, color="FFFFFF")
    sum_widths = [40, 20, 20, 18, 20, 20]
    for ci, w in enumerate(sum_widths, 1):
        ws_sum.column_dimensions[ws_sum.cell(row=1, column=ci).column_letter].width = w

    for ri, row_data in enumerate(summary_df.itertuples(index=False), 2):
        for ci, val in enumerate(row_data, 1):
            cell = ws_sum.cell(row=ri, column=ci, value=val)
            col_name = sum_cols[ci - 1]
            if col_name == col_common and val > 0:
                cell.fill = grn_fill
            elif col_name == col_not_in1 and val > 0:
                cell.fill = ora_fill
            elif col_name == col_not_in2 and val > 0:
                cell.fill = blu_fill

    # Detail sheet
    ws_det = wb.create_sheet("Critical Jobs Detail")
    det_cols   = ['Job Code', 'Job Title', 'Machinery', 'File Name', 'Match']
    det_widths = [14, 60, 40, 30, 28]
    for ci, (hdr, w) in enumerate(zip(det_cols, det_widths), 1):
        cell = ws_det.cell(row=1, column=ci, value=hdr)
        cell.fill = det_hdr_fill
        cell.font = Font(bold=True, color="FFFFFF")
        ws_det.column_dimensions[cell.column_letter].width = w

    det_cur = 2
    for _, df_det in detail_dict.items():
        for row_det in df_det.itertuples(index=False):
            match = str(row_det[4])
            rfill = grn_fill if match == 'In Both' else (
                    blu_fill if match.startswith('Only in ' + label1) else ora_fill)
            for ci, val in enumerate(row_det, 1):
                cell = ws_det.cell(row=det_cur, column=ci, value=val)
                cell.fill = rfill
                cell.alignment = wrap_aln
            det_cur += 1

    if not detail_dict:
        ws_det.cell(row=2, column=1, value="No critical jobs (Column B = 'C') found in either file.")

    out = BytesIO()
    wb.save(out)
    out.seek(0)

    return summary_df, detail_dict, out.getvalue(), label1, label2
