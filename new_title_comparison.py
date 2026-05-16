import pandas as pd
from io import BytesIO
import re
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from comparison_utils import rename_machinery


def extract_date_from_filename(filename):
    base_name = os.path.splitext(os.path.basename(filename))[0]
    date_part = base_name.split()[-1]
    try:
        if len(date_part) >= 8:
            return f"{date_part[0:2]}-{date_part[2:4]}-{date_part[4:8]}"
        return date_part
    except Exception:
        return date_part


def get_file_label(filename):
    """Return a human-friendly label from a filename, e.g. 'Harzand 15052026'.
    Splits on underscores/spaces and drops long numeric timestamp segments."""
    base = os.path.splitext(os.path.basename(filename))[0]
    parts = re.split(r'[_\s]+', base)
    parts = [p for p in parts if p and not (p.isdigit() and len(p) > 10)]
    return ' '.join(parts)


def get_vessel_name(df):
    if "Vessel" in df.columns:
        vessel_values = df["Vessel"].dropna()
        if not vessel_values.empty:
            return vessel_values.iloc[0]
    return "Unknown Vessel"


def count_titles(column):
    if column == '-' or pd.isna(column):
        return 0
    return len([x for x in column.split(', ') if x.strip()])


def prepare_excel_report(df, file1_name, file2_name, vessel1_name, vessel2_name):
    wb = Workbook()
    ws = wb.active
    ws.title = "Job Title Comparison"

    if not df.empty:
        headers = [
            'Machinery', 'Has Differences', 'Common Titles',
            'Titles only in Job List', 'Titles only in Job Status',
            'Count for Common Titles', 'Count for Job List Titles', 'Count for Job Status Titles'
        ]
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)

        for row_idx, row_data in enumerate(df.values, 2):
            row_vals = list(row_data)
            count_common = count_titles(row_data[2]) if len(row_data) > 2 else 0
            count_f1 = count_titles(row_data[3]) if len(row_data) > 3 else 0
            count_f2 = count_titles(row_data[4]) if len(row_data) > 4 else 0
            row_vals += [count_common, count_f1, count_f2]
            for col_idx, value in enumerate(row_vals, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

    fill_yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    fill_red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    fill_light_blue = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    bold_font = Font(bold=True)
    red_font = Font(color="9C0006")

    if not df.empty:
        for col in range(1, 9):
            ws.cell(row=1, column=col).font = bold_font

        for row in range(2, len(df) + 2):
            has_diff = ws.cell(row=row, column=2).value
            if has_diff == 'Yes':
                ws.cell(row=row, column=1).font = bold_font
                for col in range(1, 6):
                    header = ws.cell(row=1, column=col).value
                    if header and 'Titles only in' in header:
                        cell = ws.cell(row=row, column=col)
                        if cell.value != '-':
                            cell.fill = fill_yellow
                ws.cell(row=row, column=2).font = red_font
                ws.cell(row=row, column=2).fill = fill_red

        for col in range(1, 9):
            col_letter = chr(64 + col) if col <= 26 else chr(64 + (col - 1) // 26) + chr(65 + (col - 1) % 26)
            ws.column_dimensions[col_letter].width = 30
            for row in range(2, len(df) + 2):
                ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical='top')

    machinery_diff_sheet = wb.create_sheet(title="Machinery Differences")
    diff_machinery = df[df['Has Differences'] == 'Yes']['Machinery'].tolist() if not df.empty else []

    machinery_diff_sheet.cell(row=1, column=1, value="Machinery with Different Job Titles")
    machinery_diff_sheet.cell(row=1, column=2, value=f"Comparison: {vessel1_name} vs {vessel2_name}")
    machinery_diff_sheet.cell(row=1, column=1).font = bold_font
    machinery_diff_sheet.cell(row=1, column=2).font = bold_font
    machinery_diff_sheet.cell(row=3, column=1, value="No.")
    machinery_diff_sheet.cell(row=3, column=2, value="Machinery")
    machinery_diff_sheet.cell(row=3, column=1).font = bold_font
    machinery_diff_sheet.cell(row=3, column=2).font = bold_font

    for idx, machinery in enumerate(sorted(diff_machinery), 1):
        machinery_diff_sheet.cell(row=idx + 3, column=1, value=idx)
        machinery_diff_sheet.cell(row=idx + 3, column=2, value=machinery)
        if idx % 2 == 0:
            machinery_diff_sheet.cell(row=idx + 3, column=1).fill = fill_light_blue
            machinery_diff_sheet.cell(row=idx + 3, column=2).fill = fill_light_blue

    machinery_diff_sheet.column_dimensions['B'].width = 50

    if not diff_machinery:
        machinery_diff_sheet.cell(row=4, column=1, value="No machinery with different job titles found")
        machinery_diff_sheet.cell(row=4, column=1).font = Font(italic=True)

    try:
        output_final = BytesIO()
        wb.save(output_final)
        output_final.seek(0)
        return output_final.getvalue()
    except Exception as e:
        wb_error = Workbook()
        ws_error = wb_error.active
        ws_error.title = "Error"
        ws_error.append(["Error generating report", str(e)])
        output_error = BytesIO()
        wb_error.save(output_error)
        output_error.seek(0)
        return output_error.getvalue()


def compare_titles(file1_content, file2_content, file1_name, file2_name):
    try:
        df1 = pd.read_csv(BytesIO(file1_content))
        df2 = pd.read_csv(BytesIO(file2_content))

        label1 = get_file_label(file1_name)
        label2 = get_file_label(file2_name)
        vessel1 = get_vessel_name(df1)
        vessel2 = get_vessel_name(df2)

        first_machinery_col = None
        first_title_col = None
        second_machinery_col = None
        second_title_col = None

        if 'Machinery Location' in df1.columns:
            first_machinery_col = 'Machinery Location'
        elif 'Machinery' in df1.columns:
            first_machinery_col = 'Machinery'

        if 'Title' in df1.columns:
            first_title_col = 'Title'
        elif 'Job Title' in df1.columns:
            first_title_col = 'Job Title'
        elif 'Job Title.1' in df1.columns:
            first_title_col = 'Job Title.1'

        if 'Machinery Location' in df2.columns:
            second_machinery_col = 'Machinery Location'
        elif 'Machinery' in df2.columns:
            second_machinery_col = 'Machinery'

        if 'Job Title' in df2.columns:
            second_title_col = 'Job Title'
        elif 'Title' in df2.columns:
            second_title_col = 'Title'
        elif 'Job Title.1' in df2.columns:
            second_title_col = 'Job Title.1'

        if first_machinery_col is None:
            raise ValueError("Machinery column not found in first file. Available: " + str(df1.columns.tolist()))
        if first_title_col is None:
            raise ValueError("Title/Job Title column not found in first file. Available: " + str(df1.columns.tolist()))
        if second_machinery_col is None:
            raise ValueError("Machinery column not found in second file. Available: " + str(df2.columns.tolist()))
        if second_title_col is None:
            raise ValueError("Title/Job Title column not found in second file. Available: " + str(df2.columns.tolist()))

        df1[first_machinery_col] = df1[first_machinery_col].apply(
            lambda x: rename_machinery(str(x)) if pd.notna(x) else x
        )
        df2[second_machinery_col] = df2[second_machinery_col].apply(
            lambda x: rename_machinery(str(x)) if pd.notna(x) else x
        )

        titles_df1 = df1[[first_machinery_col, first_title_col]].copy()
        titles_df1.rename(columns={first_machinery_col: 'Machinery', first_title_col: 'Job Title'}, inplace=True)
        titles_df1.drop_duplicates(inplace=True)

        titles_df2 = df2[[second_machinery_col, second_title_col]].copy()
        titles_df2.rename(columns={second_machinery_col: 'Machinery', second_title_col: 'Job Title'}, inplace=True)
        titles_df2.drop_duplicates(inplace=True)

        titles_df1 = titles_df1[titles_df1['Machinery'].notna()]
        titles_df2 = titles_df2[titles_df2['Machinery'].notna()]

        titles_df1['Machinery'] = titles_df1['Machinery'].astype(str)
        titles_df1['Job Title'] = titles_df1['Job Title'].astype(str)
        titles_df2['Machinery'] = titles_df2['Machinery'].astype(str)
        titles_df2['Job Title'] = titles_df2['Job Title'].astype(str)

        title_comparison_results = []

        all_machinery = pd.concat([
            titles_df1['Machinery'],
            titles_df2['Machinery']
        ]).drop_duplicates().tolist()

        if label1 == label2:
            col_only1 = f'Titles only in {label1} (File 1)'
            col_only2 = f'Titles only in {label2} (File 2)'
        else:
            col_only1 = f'Titles only in {label1}'
            col_only2 = f'Titles only in {label2}'

        for machinery in all_machinery:
            if machinery == 'TOTAL':
                continue

            titles1 = titles_df1[titles_df1['Machinery'] == machinery]['Job Title'].tolist()
            titles2 = titles_df2[titles_df2['Machinery'] == machinery]['Job Title'].tolist()

            titles1 = [t for t in titles1 if t != "nan"]
            titles2 = [t for t in titles2 if t != "nan"]

            only_in_df1 = list(set(titles1) - set(titles2))
            only_in_df2 = list(set(titles2) - set(titles1))
            common_titles = list(set(titles1) & set(titles2))

            result_dict = {
                'Machinery': machinery,
                'Has Differences': 'Yes' if only_in_df1 or only_in_df2 else 'No',
                'Common Titles': ', '.join(sorted(common_titles)) if common_titles else '-',
                col_only1: ', '.join(sorted(only_in_df1)) if only_in_df1 else '-',
                col_only2: ', '.join(sorted(only_in_df2)) if only_in_df2 else '-'
            }

            if titles1 or titles2:
                title_comparison_results.append(result_dict)

        title_comparison_df = pd.DataFrame(title_comparison_results)

        if title_comparison_df.empty:
            title_comparison_df = pd.DataFrame(columns=[
                'Machinery', 'Has Differences', 'Common Titles', col_only1, col_only2
            ])
        else:
            column_order = ['Machinery', 'Has Differences', 'Common Titles']
            for col in title_comparison_df.columns:
                if 'Titles only in' in col and col not in column_order:
                    column_order.append(col)
            title_comparison_df = title_comparison_df[column_order]
            title_comparison_df.sort_values('Machinery', inplace=True)

        machinery_with_diff = title_comparison_df[
            title_comparison_df['Has Differences'] == 'Yes'
        ]['Machinery'].tolist()

        excel_data = prepare_excel_report(
            title_comparison_df, file1_name, file2_name, vessel1, vessel2
        )

        return title_comparison_df, machinery_with_diff, excel_data

    except Exception as e:
        print(f"Error in compare_titles: {str(e)}")
        empty_df = pd.DataFrame(columns=[
            'Machinery', 'Has Differences', 'Common Titles',
            'Titles only in File 1', 'Titles only in File 2'
        ])
        return empty_df, [], BytesIO().getvalue()
