import pandas as pd
import streamlit as st
from io import BytesIO
import re
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

def extract_date_from_filename(filename):
    """Extract and format date from filename."""
    base_name = os.path.splitext(os.path.basename(filename))[0]
    date_part = base_name.split()[-1]
    try:
        if len(date_part) >= 8:  # Format like 25032025
            return f"{date_part[0:2]}-{date_part[2:4]}-{date_part[4:8]}"
        else:
            return date_part
    except:
        return date_part

def get_vessel_name(df):
    """Extract vessel name from DataFrame."""
    if "Vessel" in df.columns:
        vessel_values = df["Vessel"].dropna()
        if not vessel_values.empty:
            return vessel_values.iloc[0]
    return "Unknown Vessel"

def rename_machinery(value):
    """Apply renaming rules to machinery values."""
    rename_mapping = {
        r"P1$": " P", r"Port1$": " P", r"S1$": " S", r"Starboard1$": " S", 
        r"S2$": " S", r"Starboard2$": " S", r"F$": " F", r"Forward$": " F", 
        r"A$": " A", r"Aft$": " A", r"P$": " P", r"Port$": " P", r"S$": " S", 
        r"Starboard$": " S", r"Lifeboat DavitA$": " Lifeboat Davit A",
        r"Lifeboat DavitAft$": " Lifeboat Davit A", r"LifeboatA$": " Lifeboat A",
        r"LifeboatAft$": " Lifeboat A"
    }
    
    original_value = str(value).strip()
    for pattern, replacement in rename_mapping.items():
        if re.search(pattern, original_value):
            return re.sub(pattern, replacement, original_value)
    return original_value

def compare_titles(file1_content, file2_content, file1_name, file2_name):
    """Compare job titles between two CSV files for each machinery."""
    # Read CSV files
    df_system_mgmt = pd.read_csv(BytesIO(file1_content))
    df_pms_jobs = pd.read_csv(BytesIO(file2_content))
    
    # Print column names for debugging
    print("First file columns:", df_system_mgmt.columns.tolist())
    print("Second file columns:", df_pms_jobs.columns.tolist())
    
    # Extract dates and vessel names
    date1_fmt = extract_date_from_filename(file1_name)
    date2_fmt = extract_date_from_filename(file2_name)
    vessel1 = get_vessel_name(df_system_mgmt)
    vessel2 = get_vessel_name(df_pms_jobs)
    
    # Print available columns for debugging
    print("First file columns:", df_system_mgmt.columns.tolist())
    print("Second file columns:", df_pms_jobs.columns.tolist())
    
    # Based on the sample files:
    # First file (Federal Thunderbay 25032025.csv) has columns 'Machinery Location' and 'Title'
    # Second file (Federal Thunderbay Job List 24032025.csv) has columns 'Machinery' and 'Job Title'
    
    # Determine columns based on actual file structure
    first_machinery_col = None
    first_title_col = None
    second_machinery_col = None
    second_title_col = None
    
    # Check first file
    if 'Machinery Location' in df_system_mgmt.columns:
        first_machinery_col = 'Machinery Location'
    elif 'Machinery' in df_system_mgmt.columns:
        first_machinery_col = 'Machinery'
    
    if 'Title' in df_system_mgmt.columns:
        first_title_col = 'Title'
    elif 'Job Title' in df_system_mgmt.columns:
        first_title_col = 'Job Title'
    
    # Check second file  
    if 'Machinery Location' in df_pms_jobs.columns:
        second_machinery_col = 'Machinery Location'
    elif 'Machinery' in df_pms_jobs.columns:
        second_machinery_col = 'Machinery'
    
    # Handle duplicate columns by checking if 'Job Title' appears multiple times
    col_counts = df_pms_jobs.columns.value_counts()
    
    if 'Title' in df_pms_jobs.columns:
        second_title_col = 'Title'
    elif 'Job Title' in df_pms_jobs.columns:
        second_title_col = 'Job Title'
        # Deal with duplicate 'Job Title' columns - use the last one by default
        if col_counts.get('Job Title', 0) > 1:
            print("Found duplicate 'Job Title' columns in second file")
            # Find all columns that match 'Job Title'
            job_title_cols = [col for col in df_pms_jobs.columns if col == 'Job Title']
            # Use the last one
            second_title_col = job_title_cols[-1]
        # Some files may have duplicate Job Title columns with suffixes
        if 'Job Title.1' in df_pms_jobs.columns:
            # Use the last Job Title column (sometimes the last one is the correct one)
            second_title_col = 'Job Title.1'
    
    if first_machinery_col is None:
        raise ValueError("Machinery column not found in first file. Available columns: " + 
                        str(df_system_mgmt.columns.tolist()))
    
    if first_title_col is None:
        raise ValueError("Title/Job Title column not found in first file. Available columns: " + 
                        str(df_system_mgmt.columns.tolist()))
    
    if second_machinery_col is None:
        raise ValueError("Machinery column not found in second file. Available columns: " + 
                        str(df_pms_jobs.columns.tolist()))
    
    if second_title_col is None:
        raise ValueError("Title/Job Title column not found in second file. Available columns: " + 
                        str(df_pms_jobs.columns.tolist()))
    
    print(f"Using columns: {first_machinery_col}, {first_title_col} from first file")
    print(f"Using columns: {second_machinery_col}, {second_title_col} from second file")
    
    # Print the first few rows of data for debugging
    print("\nFirst file sample data:")
    for idx, row in df_system_mgmt.head(3).iterrows():
        print(f"  Row {idx}: {first_machinery_col}={row[first_machinery_col]}, {first_title_col}={row[first_title_col]}")
    
    print("\nSecond file sample data:")
    for idx, row in df_pms_jobs.head(3).iterrows():
        print(f"  Row {idx}: {second_machinery_col}={row[second_machinery_col]}, {second_title_col}={row[second_title_col]}")
    
    # Standardize machinery names
    df_system_mgmt[first_machinery_col] = df_system_mgmt[first_machinery_col].apply(lambda x: rename_machinery(str(x)) if pd.notna(x) else x)
    df_pms_jobs[second_machinery_col] = df_pms_jobs[second_machinery_col].apply(lambda x: rename_machinery(str(x)) if pd.notna(x) else x)
    
    # Format column names for display
    col1 = f"{vessel1} ({date1_fmt})"
    col2 = f"{vessel2} ({date2_fmt})"
    
    # Prepare dataframes for title comparison
    titles_df1 = df_system_mgmt[[first_machinery_col, first_title_col]].copy()
    titles_df1.rename(columns={first_machinery_col: 'Machinery', first_title_col: 'Job Title'}, inplace=True)
    titles_df1.drop_duplicates(inplace=True)
    
    titles_df2 = df_pms_jobs[[second_machinery_col, second_title_col]].copy()
    titles_df2.rename(columns={second_machinery_col: 'Machinery', second_title_col: 'Job Title'}, inplace=True)
    titles_df2.drop_duplicates(inplace=True)
    
    # Filter out any rows with missing machinery names
    titles_df1 = titles_df1[titles_df1['Machinery'].notna()]
    titles_df2 = titles_df2[titles_df2['Machinery'].notna()]
    
    # Convert all to strings for comparison
    titles_df1['Machinery'] = titles_df1['Machinery'].astype(str)
    titles_df1['Job Title'] = titles_df1['Job Title'].astype(str)
    titles_df2['Machinery'] = titles_df2['Machinery'].astype(str)
    titles_df2['Job Title'] = titles_df2['Job Title'].astype(str)
    
    # Create a dictionary to store title comparison results
    title_comparison_results = []
    
    # Get unique machinery names from both DataFrames
    all_machinery = pd.concat([
        titles_df1['Machinery'], 
        titles_df2['Machinery']
    ]).drop_duplicates().tolist()
    
    # Compare titles for each machinery
    for machinery in all_machinery:
        if machinery == 'TOTAL':
            continue
            
        # Get titles from both DataFrames
        titles1 = titles_df1[titles_df1['Machinery'] == machinery]['Job Title'].tolist()
        titles2 = titles_df2[titles_df2['Machinery'] == machinery]['Job Title'].tolist()
        
        # Find unique titles in each dataframe
        only_in_df1 = list(set(titles1) - set(titles2))
        only_in_df2 = list(set(titles2) - set(titles1))
        common_titles = list(set(titles1) & set(titles2))
        
        # Add to results
        if only_in_df1 or only_in_df2:
            # Create a result dictionary with just the essentials first
            result_dict = {
                'Machinery': machinery,
                'Common Titles': ', '.join(sorted(common_titles)) if common_titles else '-',
                'Has Differences': 'Yes' if only_in_df1 or only_in_df2 else 'No'
            }
            
            # Add titles from first file
            if only_in_df1:
                result_dict[f'Titles only in {vessel1}'] = ', '.join(sorted(only_in_df1))
            else:
                result_dict[f'Titles only in {vessel1}'] = '-'
            
            # Add titles from second file
            if only_in_df2:
                result_dict[f'Titles only in {vessel2}'] = ', '.join(sorted(only_in_df2))
            else:
                result_dict[f'Titles only in {vessel2}'] = '-'
            
            # Add to the results list
            title_comparison_results.append(result_dict)
    
    # Create DataFrame from results
    title_comparison_df = pd.DataFrame(title_comparison_results)
    
    # Sort by machinery name
    if not title_comparison_df.empty:
        title_comparison_df.sort_values('Machinery', inplace=True)
    
    # Create Excel file in memory
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        if not title_comparison_df.empty:
            title_comparison_df.to_excel(writer, sheet_name='Job Title Comparison', index=False)
    
    # Create Excel file directly without trying to read it back
    output_final = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Job Title Comparison"
    
    if not title_comparison_df.empty:
        # Write header
        headers = title_comparison_df.columns.tolist()
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)
        
        # Write data - more cautious approach
        for row_idx, row_data in enumerate(title_comparison_df.values, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Define styles
    fill_yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Light yellow
    fill_red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red
    bold_font = Font(bold=True)
    red_font = Font(color="9C0006")  # Dark red
    
    # Apply styling if there's a Job Title Comparison sheet
    if 'Job Title Comparison' in wb.sheetnames:
        ws = wb['Job Title Comparison']
        
        # Format headers
        for col in range(1, len(title_comparison_df.columns) + 1):
            ws.cell(row=1, column=col).font = bold_font
        
        # Format data rows
        for row in range(2, len(title_comparison_df) + 2):
            # Get the "Has Differences" value
            has_diff = ws.cell(row=row, column=5).value
            
            if has_diff == 'Yes':
                # Highlight machinery name
                ws.cell(row=row, column=1).font = bold_font
                
                # Highlight title columns
                titles_only_in_1 = ws.cell(row=row, column=2)
                titles_only_in_2 = ws.cell(row=row, column=3)
                
                if titles_only_in_1.value != '-':
                    titles_only_in_1.fill = fill_yellow
                
                if titles_only_in_2.value != '-':
                    titles_only_in_2.fill = fill_yellow
                
                # Highlight "Has Differences" column
                ws.cell(row=row, column=5).font = red_font
                ws.cell(row=row, column=5).fill = fill_red
        
        # Adjust column widths and set text wrapping
        for col in range(1, len(title_comparison_df.columns) + 1):
            col_letter = chr(64 + col)  # Convert column number to letter (A, B, C, ...)
            ws.column_dimensions[col_letter].width = 30
            
            # Set text wrapping for all cells
            for row in range(2, len(title_comparison_df) + 2):
                cell = ws.cell(row=row, column=col)
                cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Save the styled workbook
    output_final = BytesIO()
    wb.save(output_final)
    output_final.seek(0)
    
    return title_comparison_df, output_final.getvalue()

def render_title_comparison_app():
    """Render the Streamlit app for job title comparison."""
    # Page config is now set in main.py

    st.title("🚢 Job Title Comparison Tool")

    st.markdown("""
    This tool compares job titles for machinery between two CSV files.

    **Instructions:**
    1. Upload two CSV files containing machinery data
    2. The files should have machinery and job title columns:
       - Acceptable machinery column names include: 'Machinery', 'Machinery Location'
       - Acceptable job title column names include: 'Job Title', 'Title'
       - The tool will automatically detect and use the correct columns
    3. The tool will analyze the data and show machinery with different job titles
    4. Download the highlighted Excel report for offline use
    """)

    # File upload section
    col1, col2 = st.columns(2)

    with col1:
        st.subheader("First CSV File (System Management)")
        file1 = st.file_uploader("Upload System Management CSV", type=['csv'], key="file1")
        if file1:
            df1 = pd.read_csv(file1)
            st.write("Preview of first file:")
            st.dataframe(df1.head(), use_container_width=True)
            st.info(f"Total rows: {len(df1)}")
            
            # Check for machinery columns
            has_machinery_col = False
            for col_name in ['Machinery', 'Machinery Location']:
                if col_name in df1.columns:
                    has_machinery_col = True
                    break
                    
            # Check for title columns
            has_title_col = False
            for col_name in ['Title', 'Job Title']:
                if col_name in df1.columns:
                    has_title_col = True
                    break
            
            if has_machinery_col and has_title_col:
                st.success("✅ Required column types found!")
            else:
                missing = []
                if not has_machinery_col:
                    missing.append("Machinery column ('Machinery' or 'Machinery Location')")
                if not has_title_col:
                    missing.append("Title column ('Job Title' or 'Title')")
                st.error(f"❌ Missing required columns: {', '.join(missing)}")
                st.write("Available columns:", df1.columns.tolist())

    with col2:
        st.subheader("Second CSV File (PMS Jobs)")
        file2 = st.file_uploader("Upload PMS Jobs CSV", type=['csv'], key="file2")
        if file2:
            df2 = pd.read_csv(file2)
            st.write("Preview of second file:")
            st.dataframe(df2.head(), use_container_width=True)
            st.info(f"Total rows: {len(df2)}")
            
            # Check for machinery columns
            has_machinery_col = False
            for col_name in ['Machinery', 'Machinery Location']:
                if col_name in df2.columns:
                    has_machinery_col = True
                    break
                    
            # Check for title columns
            has_title_col = False
            for col_name in ['Title', 'Job Title']:
                if col_name in df2.columns:
                    has_title_col = True
                    break
            
            if has_machinery_col and has_title_col:
                st.success("✅ Required column types found!")
            else:
                missing = []
                if not has_machinery_col:
                    missing.append("Machinery column ('Machinery' or 'Machinery Location')")
                if not has_title_col:
                    missing.append("Title column ('Job Title' or 'Title')")
                st.error(f"❌ Missing required columns: {', '.join(missing)}")
                st.write("Available columns:", df2.columns.tolist())

    if file1 and file2:
        try:
            # Process the files
            with st.spinner('Comparing job titles between files...'):
                title_comparison_df, excel_data = compare_titles(
                    file1.getvalue(),
                    file2.getvalue(),
                    file1.name,
                    file2.name
                )
            
            # Job Title Comparison section
            st.subheader("📋 Job Title Comparison")
            
            if not title_comparison_df.empty:
                # Filter to only show items with differences
                title_diff_df = title_comparison_df[title_comparison_df['Has Differences'] == 'Yes']
                
                if not title_diff_df.empty:
                    st.warning(f"Found {len(title_diff_df)} machinery items with different job titles:")
                    
                    # 1. Simple List of Machinery with Different Titles
                    st.subheader("📋 Machinery List with Different Job Titles")
                    
                    # Create a clean list for display
                    machinery_list = sorted(title_diff_df['Machinery'].tolist())
                    machinery_display = ""
                    for i, machinery in enumerate(machinery_list):
                        machinery_display += f"{i+1}. {machinery}\n"
                    
                    st.code(machinery_display, language=None)
                    
                    # 2. Display a detailed comparison table
                    st.subheader("🔄 Detailed Job Title Comparison by File")

                    # Get column names for titles in each file (different approach)
                    # We're looking for just one 'Titles only in' column since that's what we're seeing
                    titles_only_col = next((col for col in title_diff_df.columns if 'Titles only in' in col), None)
                    
                    # If we found at least one column
                    if titles_only_col is not None:
                        # We need to handle the case where we only have one vessel's titles
                        file1_col = titles_only_col
                        
                        # Create a fallback column name for the second file based on the first
                        current_vessel = titles_only_col.replace('Titles only in ', '')
                        other_vessel = "second file" if "first" in current_vessel.lower() else "first file"
                        
                        # Check if there's a second "Titles only in" column
                        titles_only_cols = [col for col in title_diff_df.columns if 'Titles only in' in col]
                        if len(titles_only_cols) > 1:
                            file2_col = titles_only_cols[1]
                        else:
                            # Try to find the column name for the second file
                            file2_col = None
                            for col in title_diff_df.columns:
                                if col not in ['Machinery', 'Common Titles', 'Has Differences', file1_col]:
                                    file2_col = col
                                    break
                            
                            # If we still can't find it, use a fallback approach
                            if file2_col is None:
                                file2_col = "Titles from Second File"
                                # Create a fallback column
                                title_diff_df[file2_col] = '-'
                        
                        # Extract vessel names for cleaner display
                        vessel1_name = file1_col.replace('Titles only in ', '')
                        vessel2_name = file2_col.replace('Titles only in ', '')
                        
                        # Create comparison data
                        comparison_data = []
                        
                        for _, row in title_diff_df.iterrows():
                            machinery = row['Machinery']
                            titles_in_file1 = row[file1_col] if row[file1_col] != '-' else 'None'
                            titles_in_file2 = row[file2_col] if row[file2_col] != '-' else 'None'
                            
                            comparison_data.append({
                                "Machinery": machinery,
                                f"Job Titles in {vessel1_name}": titles_in_file1,
                                f"Job Titles in {vessel2_name}": titles_in_file2
                            })
                        
                        # Create and display comparison DataFrame
                        comparison_df = pd.DataFrame(comparison_data)
                        st.dataframe(comparison_df, use_container_width=True)
                        
                        # 3. Display a detailed text list for easy copying
                        st.subheader("📝 Text Comparison (Job Title Differences)")
                        
                        detailed_list = []
                        for _, row in title_diff_df.iterrows():
                            machinery = row['Machinery']
                            titles_in_file1 = row[file1_col] if row[file1_col] != '-' else 'None'
                            titles_in_file2 = row[file2_col] if row[file2_col] != '-' else 'None'
                            
                            detailed_list.append(f"MACHINERY: {machinery}")
                            detailed_list.append(f"- {vessel1_name}: {titles_in_file1}")
                            detailed_list.append(f"- {vessel2_name}: {titles_in_file2}")
                            detailed_list.append("")
                        
                        detailed_text = "\n".join(detailed_list)
                        st.text_area("Copy-paste friendly comparison:", detailed_text, height=300)
                    else:
                        st.error("Could not find title comparison columns in the data. Please check the CSV files.")
                        
                        # Display all columns for debugging
                        st.subheader("Debugging: Available Columns")
                        st.json(title_diff_df.columns.tolist())
                        
                        # Display the raw data for inspection
                        st.subheader("Debugging: Raw Comparison Data")
                        st.dataframe(title_diff_df)
                    
                    # Custom styling function for title comparison
                    def highlight_title_differences(row):
                        # Create a style list with the exact length of the row
                        column_count = len(title_diff_df.columns)
                        styles = [''] * column_count
                        
                        # Highlight machinery column (first column)
                        styles[0] = 'font-weight: bold'
                        
                        # Get column indices for title columns
                        # These are typically the second and third columns (indices 1 and 2)
                        vessel1_idx = 1 if column_count > 1 else 0
                        vessel2_idx = 2 if column_count > 2 else 0
                        
                        # Get column names
                        vessel1_title_col = title_diff_df.columns[vessel1_idx] if vessel1_idx > 0 else None
                        vessel2_title_col = title_diff_df.columns[vessel2_idx] if vessel2_idx > 0 else None
                        
                        # Highlight title columns if they exist
                        if vessel1_title_col and row[vessel1_title_col] != '-':
                            styles[vessel1_idx] = 'background-color: #FFEB9C'  # Light yellow
                            
                        if vessel2_title_col and row[vessel2_title_col] != '-':
                            styles[vessel2_idx] = 'background-color: #FFEB9C'  # Light yellow
                        
                        # Find the "Has Differences" column - it should be the last column or column 4
                        diff_idx = column_count - 1  # Default to last column
                        if 'Has Differences' in row:
                            for i, col_name in enumerate(title_diff_df.columns):
                                if col_name == 'Has Differences':
                                    diff_idx = i
                                    break
                            
                            # Highlight "Has Differences" column
                            if row['Has Differences'] == 'Yes':
                                styles[diff_idx] = 'background-color: #FFC7CE; color: #9C0006; font-weight: bold'  # Red
                        
                        return styles
                    
                    st.dataframe(
                        title_diff_df.style.apply(highlight_title_differences, axis=1),
                        use_container_width=True
                    )
                    
                    # Show raw data for inspection
                    st.subheader("🔎 Examples of Job Title Differences")
                    
                    # Sample up to 5 machinery items to show detailed differences
                    sample_count = min(5, len(title_diff_df))
                    if sample_count > 0:
                        st.write("Below are examples of machinery with different job titles:")
                        
                        sample_machines = title_diff_df['Machinery'].sample(sample_count).tolist()
                        
                        for idx, machinery in enumerate(sample_machines):
                            row = title_diff_df[title_diff_df['Machinery'] == machinery].iloc[0]
                            
                            st.write(f"**{idx+1}. {machinery}**")
                            
                            col1, col2 = st.columns(2)
                            
                            with col1:
                                # Instead of using specific column names, let's display first non-Machinery column
                                first_col = [col for col in title_diff_df.columns if col != 'Machinery' and col != 'Has Differences' and col != 'Common Titles'][0]
                                st.write(f"**{first_col}:**")
                                if row[first_col] != '-':
                                    st.write(f"{row[first_col]}")
                                else:
                                    st.write("*None*")
                            
                            with col2:
                                # Display Common Titles column
                                st.write(f"**Common Titles:**")
                                if 'Common Titles' in row and row['Common Titles'] != '-':
                                    st.write(f"{row['Common Titles']}")
                                else:
                                    st.write("*None*")
                            
                            st.write("---")
                else:
                    st.success("No job title differences found for any machinery!")
            else:
                st.info("No job title comparison data generated. Please check if both files have matching machinery.")
            
            # Download section
            st.subheader("📥 Download Report")
            st.write("Download the detailed Excel report with highlighted job title differences:")
            st.download_button(
                label="Download Job Title Comparison Report",
                data=excel_data,
                file_name="Job_Title_Comparison.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                help="Click to download the job title comparison report in Excel format"
            )
            
        except Exception as e:
            st.error(f"Error comparing job titles: {str(e)}")
            st.error("Please ensure both files have the required columns and proper format:")
            st.info("Acceptable column names: 'Machinery'/'Machinery Location' for machinery data")
            st.info("Acceptable column names: 'Job Title'/'Title' for job title data")
            st.info("The tool will automatically detect and use the appropriate columns if available")
    else:
        st.info("👆 Please upload both CSV files to generate the title comparison report")

    # Add footer
    st.markdown("---")
    st.markdown("Built with ❤️ using Streamlit")

if __name__ == "__main__":
    render_title_comparison_app()